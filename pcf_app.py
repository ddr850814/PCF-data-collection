"""PCF 收集器 - 主 GUI 应用程序（Python 版本）

功能等同于 VB.NET 版 PCF 收集器：
1. 选择文件夹，批量解析 PCF 文件并存入 SQLite 数据库
2. 打开已有数据库，浏览组件数据
3. 右键导出数据到 Excel
"""

import os
import re
import threading
import traceback
import tkinter as tk
from tkinter import filedialog, messagebox

import ttkbootstrap as ttkb
from ttkbootstrap.constants import *

from database_helper import DatabaseHelper
from pcf_data_collector import PcfDataCollector
from pcf_parser import PcfParser

try:
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class PcfApp:
    """PCF 收集器主应用程序"""

    COMPONENT_TYPES = PcfDataCollector.COMPONENT_TYPES

    def __init__(self, root):
        self.root = root
        self.root.title("PCF收集")
        self.root.geometry("1000x650")

        # 状态变量
        self.db_helper = None
        self.data_collector = None
        self.db_path = None
        self.current_columns = []
        self.current_rows = []
        self.success_count = 0
        self.error_count = 0
        self.error_log = []
        self._processing = False

        self._build_ui()

    def _build_ui(self):
        """构建界面"""
        # 配置 Treeview 样式 - 模拟网格线效果
        style = ttkb.Style.get_instance()
        if style:
            style.configure("Treeview.Heading", relief="solid", borderwidth=1, font=("-size", 10, "-weight", "bold"))
            style.configure("Treeview", rowheight=26)
            # 网格线样式
            style.layout("Grid.Treeview", style.layout("Treeview"))
            style.configure("Grid.Treeview", rowheight=26)
            style.configure("Grid.Treeview.Heading", relief="solid", borderwidth=1,
                            font=("-size", 10, "-weight", "bold"))

        # 主容器：左右分割
        main_paned = ttkb.Panedwindow(self.root, orient=tk.HORIZONTAL, bootstyle=SECONDARY)
        main_paned.pack(fill=BOTH, expand=YES, padx=6, pady=6)

        # ========== 左侧：数据表格 ==========
        left_frame = ttkb.Frame(main_paned)
        main_paned.add(left_frame, weight=2)

        # 数据表格标题
        ttkb.Label(left_frame, text="数据列表", font=("-size", 11, "-weight", "bold"),
                   bootstyle=INVERSE).pack(fill=X, pady=(0, 4))

        # DataGridView 等价物 - Treeview
        tree_container = ttkb.Frame(left_frame)
        tree_container.pack(fill=BOTH, expand=YES)

        self.tree = ttkb.Treeview(tree_container, show="headings", bootstyle=PRIMARY, style="Grid.Treeview")
        tree_scroll_y = ttkb.Scrollbar(tree_container, orient=VERTICAL, command=self.tree.yview, bootstyle=ROUND)
        tree_scroll_x = ttkb.Scrollbar(tree_container, orient=HORIZONTAL, command=self.tree.xview, bootstyle=ROUND)
        self.tree.configure(yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)

        self.tree.grid(row=0, column=0, sticky=NSEW)
        tree_scroll_y.grid(row=0, column=1, sticky=NS)
        tree_scroll_x.grid(row=1, column=0, sticky=EW)
        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)

        # 右键菜单
        self.context_menu = tk.Menu(self.root, tearoff=0)
        self.context_menu.add_command(label="导出 Excel", command=self._on_export)
        self.tree.bind("<Button-3>", self._on_right_click)

        # 选中行变化事件
        self.tree.bind("<<TreeviewSelect>>", self._on_cell_changed)

        # ========== 右侧：控制面板 + 属性面板 ==========
        right_frame = ttkb.Frame(main_paned)
        main_paned.add(right_frame, weight=1)
        self._main_paned = main_paned
        self.root.after(50, lambda: self._main_paned.sashpos(0, 600))

        # 按钮行
        btn_frame = ttkb.Labelframe(right_frame, text="操作", padding=8)
        btn_frame.pack(fill=X, padx=4, pady=4)

        self.btn_new = ttkb.Button(btn_frame, text="新建数据库", command=self._on_new_database,
                                   bootstyle=(SUCCESS, OUTLINE))
        self.btn_new.pack(side=LEFT, padx=4)

        self.btn_open = ttkb.Button(btn_frame, text="打开数据库", command=self._on_open_database,
                                    bootstyle=(INFO, OUTLINE))
        self.btn_open.pack(side=LEFT, padx=4)

        # 编码选择（紧跟在打开数据库按钮右侧）
        ttkb.Label(btn_frame, text="编码:", bootstyle=PRIMARY).pack(side=LEFT, padx=(8, 4))
        self.encoding_options = [
            ("自动（GBK 优先）", "auto"),
            ("GBK", "gbk"),
            ("UTF-8", "utf-8"),
            ("UTF-8-SIG", "utf-8-sig"),
            ("Latin-1", "latin-1"),
        ]
        self.enc_var = tk.StringVar(value="自动（GBK 优先）")
        self.combo_encoding = ttkb.Combobox(
            btn_frame, state="readonly",
            values=[label for label, _ in self.encoding_options],
            textvariable=self.enc_var, bootstyle=PRIMARY)
        self.combo_encoding.current(0)
        self.combo_encoding.pack(side=LEFT, fill=X, expand=YES)

        # 下拉框
        combo_frame = ttkb.Frame(right_frame, padding=(8, 4))
        combo_frame.pack(fill=X)

        ttkb.Label(combo_frame, text="类型:", bootstyle=PRIMARY).pack(side=LEFT, padx=(0, 4))
        self.combo_type = ttkb.Combobox(combo_frame, state="readonly", values=["COUNT"], bootstyle=PRIMARY)
        self.combo_type.current(0)
        self.combo_type.state(["disabled"])
        self.combo_type.pack(side=LEFT, fill=X, expand=YES)
        self.combo_type.bind("<<ComboboxSelected>>", self._on_combo_changed)

        # PropertyGrid 等价物 - 属性 Treeview
        prop_labelframe = ttkb.Labelframe(right_frame, text="属性详情", padding=4)
        prop_labelframe.pack(fill=BOTH, expand=YES, padx=4, pady=4)

        self.prop_tree = ttkb.Treeview(prop_labelframe, columns=("value",), show="tree headings",
                                       bootstyle=INFO, style="Grid.Treeview")
        prop_scroll = ttkb.Scrollbar(prop_labelframe, orient=VERTICAL, command=self.prop_tree.yview,
                                     bootstyle=ROUND)
        self.prop_tree.configure(yscrollcommand=prop_scroll.set)
        self.prop_tree.heading("#0", text="属性")
        self.prop_tree.heading("value", text="值")
        self.prop_tree.column("#0", width=150)
        self.prop_tree.column("value", width=200)
        self.prop_tree.pack(side=LEFT, fill=BOTH, expand=YES)
        prop_scroll.pack(side=RIGHT, fill=Y)

        # 状态栏
        self.status_var = tk.StringVar()
        self.status_var.set("就绪")
        status_bar = ttkb.Label(self.root, textvariable=self.status_var, bootstyle=(INVERSE, SECONDARY),
                                anchor=W, padding=(8, 4))
        status_bar.pack(fill=X, side=BOTTOM)

    # ==================== 事件处理 ====================

    def _on_right_click(self, event):
        """右键菜单"""
        if self.db_helper is not None and self.db_path is not None:
            self.context_menu.tk_popup(event.x_root, event.y_root)

    def _on_new_database(self):
        """新建数据库 - 选择文件夹并处理 PCF 文件"""
        folder = filedialog.askdirectory(title="选择包含 PCF 文件的文件夹")
        if not folder:
            return

        pcf_files = []
        for root_dir, dirs, files in os.walk(folder):
            for f in files:
                if f.lower().endswith(".pcf"):
                    pcf_files.append(os.path.join(root_dir, f))

        if not pcf_files:
            messagebox.showwarning("提示", "木有发现pcf")
            return

        self.db_path = os.path.join(folder, "pcf.db")
        self._processing = True
        self.btn_new.config(state=tk.DISABLED)
        self.btn_open.config(state=tk.DISABLED)

        # 关闭当前数据库连接（避免覆写时文件被占用）
        if self.db_helper is not None:
            self.db_helper.close()
            self.db_helper = None

        # 后台线程处理
        thread = threading.Thread(target=self._process_files, args=(folder, pcf_files), daemon=True)
        thread.start()

    def _process_files(self, folder, pcf_files):
        """后台处理 PCF 文件（在工作线程中运行）"""
        total = len(pcf_files)
        self.success_count = 0
        self.error_count = 0
        self.error_log.clear()

        # 初始化数据库
        if os.path.exists(self.db_path):
            os.remove(self.db_path)

        self.db_helper = DatabaseHelper(self.db_path)

        # 创建数据收集器
        self.data_collector = PcfDataCollector()
        self.data_collector.db_helper = self.db_helper

        # 创建所有表
        self.data_collector.create_all_tables()

        # 解析所有 PCF 文件
        encoding = "auto"
        selected_label = self.enc_var.get()
        for label, value in self.encoding_options:
            if label == selected_label:
                encoding = value
                break
        parser = PcfParser(self.data_collector, encoding=encoding)
        fid = 0
        for i, pcf_file in enumerate(pcf_files):
            try:
                consumed = parser.process_single_file(pcf_file, fid)
                fid += consumed
                self.success_count += 1
            except Exception as e:
                self.error_count += 1
                error_msg = f"文件 {i + 1}: {os.path.basename(pcf_file)} - {e}"
                self.error_log.append(error_msg)

            # 更新进度
            current = i + 1
            self.root.after(0, lambda c=current, t=total: self.status_var.set(
                f"{c}/{t} (成功:{self.success_count}, 失败:{self.error_count})"))

        # 插入数据 & 创建视图
        try:
            self.root.after(0, lambda: self.status_var.set("插入数据表..."))
            self.data_collector.insert_all_data()

            self.root.after(0, lambda: self.status_var.set("插入视图..."))
            self.data_collector.create_sqlite_view_all(self.COMPONENT_TYPES)

            # COUNT 视图
            self.db_helper.execute_non_query("""
                DROP VIEW IF EXISTS COUNT;
                CREATE VIEW IF NOT EXISTS COUNT AS
                SELECT PIPELINE.FILENAME, REFERENCE, TOTAL.ITEMCODE, ITEMDESCRIPTION, COUNT(*) As COUNT
                FROM (SELECT * FROM TOTAL GROUP BY UCI,FILENAME) AS TOTAL
                Join PIPELINE ON PIPELINE.FILENAME = TOTAL.FILENAME
                Group BY REFERENCE, TOTAL.ITEMCODE, ITEMDESCRIPTION
                ORDER BY REFERENCE
            """)

            # 各组件类型视图
            self._create_component_views()

            # 带属性的组件视图
            self.data_collector.create_component_with_attributes_view()
        except Exception:
            err_msg = traceback.format_exc()
            self.root.after(0, lambda: self._on_processing_error(err_msg))
            return

        # 完成
        self.root.after(0, self._on_processing_done)

    def _create_component_views(self):
        """为每种组件类型创建视图"""
        for comp_type in self.COMPONENT_TYPES:
            sql = self._get_view_sql(comp_type)
            if sql:
                self.db_helper.execute_non_query(sql)

    def _get_view_sql(self, comp_type):
        """获取组件视图 SQL"""
        xname = f"X{comp_type}"
        if comp_type == "PIPE":
            return f"""DROP VIEW IF EXISTS {xname};
                CREATE VIEW IF NOT EXISTS {xname} AS
                SELECT PIPELINE.FILENAME,UCI, REFERENCE, ITEMCODE, ITEMDESCRIPTION, CUTPIECELENGTH
                FROM (SELECT * FROM {comp_type} GROUP BY UCI,FILENAME) AS {comp_type}
                JOIN PIPELINE ON PIPELINE.FILENAME = {comp_type}.FILENAME
                ORDER BY REFERENCE"""
        elif comp_type == "BEND":
            return """DROP VIEW IF EXISTS XBEND;
                CREATE VIEW IF NOT EXISTS XBEND AS
                SELECT PIPELINE.FILENAME,UCI, REFERENCE,ITEMCODE, ITEMDESCRIPTION,CUTPIECELENGTH,ANGLE
                FROM (SELECT * FROM BEND GROUP BY UCI,FILENAME) AS BEND
                JOIN PIPELINE ON PIPELINE.FILENAME = BEND.FILENAME
                ORDER BY REFERENCE"""
        elif comp_type in ("INSTRUMENT", "INSTRUMENT_ANGLE"):
            return f"""DROP VIEW IF EXISTS {xname};
                CREATE VIEW IF NOT EXISTS {xname} AS
                SELECT PIPELINE.FILENAME,UCI,REFERENCE,ITEMCODE, TAG
                FROM (SELECT * FROM {comp_type} GROUP BY UCI,FILENAME) AS {comp_type}
                JOIN PIPELINE ON PIPELINE.FILENAME = {comp_type}.FILENAME
                ORDER BY REFERENCE"""
        elif comp_type == "INSTRUMENT_RETURN":
            return """DROP VIEW IF EXISTS XINSTRUMENT_RETURN;
                CREATE VIEW IF NOT EXISTS XINSTRUMENT_RETURN AS
                SELECT PIPELINE.FILENAME,UCI,REFERENCE,ITEMCODE, TAG, ITEMDESCRIPTION
                FROM (SELECT * FROM INSTRUMENT_RETURN GROUP BY UCI,FILENAME) AS INSTRUMENT_RETURN
                JOIN PIPELINE ON PIPELINE.FILENAME = INSTRUMENT_RETURN.FILENAME
                ORDER BY REFERENCE"""
        elif comp_type == "TRAP_OFFSET":
            return """DROP VIEW IF EXISTS XTRAP_OFFSET;
                CREATE VIEW IF NOT EXISTS XTRAP_OFFSET AS
                SELECT PIPELINE.FILENAME,UCI,REFERENCE,ITEMCODE, TAG, ITEMDESCRIPTION
                FROM (SELECT * FROM TRAP_OFFSET GROUP BY UCI,FILENAME) AS TRAP_OFFSET
                JOIN PIPELINE ON PIPELINE.FILENAME = TRAP_OFFSET.FILENAME
                ORDER BY REFERENCE"""
        elif comp_type == "SUPPORT":
            return """DROP VIEW IF EXISTS XSUPPORT;
                CREATE VIEW IF NOT EXISTS XSUPPORT AS
                SELECT PIPELINE.FILENAME,UCI,REFERENCE,NAME, ITEMDESCRIPTION,SUPPORTDIRECTION,SUPPORTTYPE
                FROM (SELECT * FROM SUPPORT GROUP BY UCI,FILENAME) AS SUPPORT
                JOIN PIPELINE ON PIPELINE.FILENAME = SUPPORT.FILENAME
                ORDER BY REFERENCE"""
        elif comp_type == "GASKET":
            return """DROP VIEW IF EXISTS XGASKET;
                CREATE VIEW IF NOT EXISTS XGASKET AS
                SELECT PIPELINE.FILENAME,GASKET.UCI,REFERENCE,GASKET.ITEMCODE, GASKET.ITEMDESCRIPTION,
                       TOTAL.ITEMCODE,TOTAL.ITEMDESCRIPTION
                FROM (SELECT * FROM GASKET GROUP BY UCI,FILENAME) AS GASKET
                JOIN PIPELINE ON PIPELINE.FILENAME = GASKET.FILENAME
                JOIN TOTAL ON TOTAL.FILENAME = GASKET.FILENAME AND TOTAL.COMPONENTIDENTIFIER=GASKET.MASTERCOMPONENTIDENTIFIER
                ORDER BY REFERENCE"""
        elif comp_type == "BOLT":
            return """DROP VIEW IF EXISTS XBOLT;
                CREATE VIEW IF NOT EXISTS XBOLT AS
                SELECT PIPELINE.FILENAME,BOLT.UCI,REFERENCE,BOLTITEMCODE,BOLTITEMDESCRIPTION,
                       BOLTDIA, BOLTQUANTITY, BOLTLENGTH, ITEMCODE,ITEMDESCRIPTION
                FROM (SELECT * FROM BOLT GROUP BY UCI,FILENAME) AS BOLT
                JOIN PIPELINE ON PIPELINE.FILENAME = BOLT.FILENAME
                JOIN TOTAL ON TOTAL.FILENAME = BOLT.FILENAME AND TOTAL.COMPONENTIDENTIFIER=BOLT.MASTERCOMPONENTIDENTIFIER
                ORDER BY REFERENCE"""
        elif comp_type == "WELD":
            return """DROP VIEW IF EXISTS XWELD;
                CREATE VIEW IF NOT EXISTS XWELD AS
                SELECT PIPELINE.FILENAME, WELD.UCI, REFERENCE, Q.NPD,
                       TOTAL.ITEMCODE, TOTAL.ITEMDESCRIPTION
                FROM (SELECT * FROM WELD GROUP BY UCI, FILENAME) AS WELD
                JOIN PIPELINE ON PIPELINE.FILENAME = WELD.FILENAME
                JOIN TOTAL ON TOTAL.FILENAME = WELD.FILENAME AND TOTAL.COMPONENTIDENTIFIER = WELD.MASTERCOMPONENTIDENTIFIER
                JOIN (SELECT * FROM END_POINT GROUP BY FILENAME, UCI) AS Q ON Q.FILENAME = WELD.FILENAME AND Q.UCI = WELD.UCI
                ORDER BY REFERENCE"""
        elif comp_type == "TEE_STUB":
            return """DROP VIEW IF EXISTS XTEE_STUB;
                CREATE VIEW IF NOT EXISTS XTEE_STUB AS
                SELECT PIPELINE.FILENAME,UCI,REFERENCE
                FROM (SELECT * FROM TEE_STUB GROUP BY UCI,FILENAME) AS TEE_STUB
                JOIN PIPELINE ON PIPELINE.FILENAME = TEE_STUB.FILENAME
                ORDER BY REFERENCE"""
        elif comp_type == "_UNION":
            return """DROP VIEW IF EXISTS X_UNION;
                CREATE VIEW IF NOT EXISTS X_UNION AS
                SELECT PIPELINE.FILENAME,UCI,REFERENCE,ITEMCODE, ITEMDESCRIPTION
                FROM (SELECT * FROM _UNION GROUP BY UCI,FILENAME) AS _UNION
                JOIN PIPELINE ON PIPELINE.FILENAME = _UNION.FILENAME
                ORDER BY REFERENCE"""
        elif comp_type == "MESSAGE":
            return """DROP VIEW IF EXISTS XMESSAGE;
                CREATE VIEW IF NOT EXISTS XMESSAGE AS
                SELECT PIPELINE.FILENAME,UCI,REFERENCE,TEXT
                FROM (SELECT * FROM MESSAGE GROUP BY UCI,FILENAME) AS MESSAGE
                JOIN PIPELINE ON PIPELINE.FILENAME = MESSAGE.FILENAME
                ORDER BY REFERENCE"""
        elif comp_type == "VALVE":
            return """DROP VIEW IF EXISTS XVALVE;
                CREATE VIEW IF NOT EXISTS XVALVE AS
                SELECT PIPELINE.FILENAME,UCI,REFERENCE,ITEMCODE, ITEMDESCRIPTION,SPINDLEDIRECTION,DIRECTION
                FROM (SELECT * FROM VALVE GROUP BY UCI,FILENAME) AS VALVE
                JOIN PIPELINE ON PIPELINE.FILENAME = VALVE.FILENAME
                ORDER BY REFERENCE"""
        else:
            return f"""DROP VIEW IF EXISTS {xname};
                CREATE VIEW IF NOT EXISTS {xname} AS
                SELECT PIPELINE.FILENAME,UCI,REFERENCE,ITEMCODE, ITEMDESCRIPTION
                FROM (SELECT * FROM {comp_type} GROUP BY UCI,FILENAME) AS {comp_type}
                JOIN PIPELINE ON PIPELINE.FILENAME = {comp_type}.FILENAME
                ORDER BY REFERENCE"""

    def _on_processing_done(self):
        """文件处理完成回调"""
        self._processing = False
        self.btn_new.config(state=tk.NORMAL)
        self.btn_open.config(state=tk.NORMAL)
        self.status_var.set("写入数据库")

        # 显示处理结果
        if self.error_count > 0:
            error_summary = f"处理完成。成功: {self.success_count}, 失败: {self.error_count}\n\n"
            error_summary += "错误详情 (前10条):\n"
            error_summary += "\n".join(self.error_log[:10])
            if len(self.error_log) > 10:
                error_summary += f"\n... 还有 {len(self.error_log) - 10} 个错误"
            messagebox.showwarning("PCF处理结果", error_summary)
        elif self.success_count > 0:
            messagebox.showinfo("PCF处理完成", f"所有 {self.success_count} 个文件处理成功！")

        self._populate_combo_and_load()

    def _on_processing_error(self, err_msg):
        """处理过程中出错回调（显示真实异常，避免线程静默退出）"""
        self._processing = False
        self.btn_new.config(state=tk.NORMAL)
        self.btn_open.config(state=tk.NORMAL)
        self.status_var.set("处理出错")
        messagebox.showerror("处理出错", err_msg[-2000:])

    def _on_open_database(self):
        """打开已有数据库"""
        db_file = filedialog.askopenfilename(title="选择数据库文件", filetypes=[("SQLite", "*.db")])
        if not db_file:
            return

        self.db_path = db_file
        self.db_helper = DatabaseHelper(db_file)
        self._populate_combo_and_load()

    def _populate_combo_and_load(self):
        """动态填充下拉框并加载数据"""
        # 获取有数据的组件类型
        values = ["COUNT"]
        for comp_type in self.COMPONENT_TYPES:
            view_name = "X_UNION" if comp_type == "_UNION" else f"X{comp_type}"
            try:
                count = self.db_helper.execute_scalar(f"SELECT COUNT(*) FROM {view_name}")
                if count and count > 0:
                    display_name = comp_type.replace("_", "-")
                    values.append(display_name)
            except Exception:
                continue

        self.combo_type.state(["!disabled"])
        self.combo_type["values"] = values
        if self.combo_type.get() not in values:
            self.combo_type.current(0)
        self._load_data()

    # ==================== 数据加载 ====================

    def _on_combo_changed(self, event):
        self._load_data()

    def _load_data(self):
        """根据选择加载表格数据"""
        selected = self.combo_type.get()
        if not selected or self.db_helper is None:
            return

        if selected == "COUNT":
            sql = "SELECT * FROM COUNT;"
        else:
            db_type = selected.replace("-", "_")
            if db_type == "UNION":
                db_type = "_UNION"
            view_name = f"X{db_type}"
            sql = f"SELECT * FROM {view_name};"

        try:
            columns, rows = self.db_helper.fill_data_table(sql)
        except Exception as e:
            messagebox.showerror("错误", str(e))
            return

        self.current_columns = columns
        self.current_rows = rows

        # 更新 Treeview（COUNT隐藏FILENAME，其他类型隐藏FILENAME和UCI）
        self._update_treeview(columns, rows, hide_first_cols=(1 if selected == "COUNT" else 2))

        # 选中第一行（触发选中事件加载属性）
        children = self.tree.get_children()
        if children:
            self.tree.selection_set(children[0])
            self.tree.focus(children[0])

    def _update_treeview(self, columns, rows, hide_first_cols=0):
        """更新 Treeview 数据。hide_first_cols 为要隐藏的前N列数量。"""
        self.tree.delete(*self.tree.get_children())

        # 先清 displaycolumns 再清 columns，避免 Tk 校验旧列名报错
        self.tree.configure(displaycolumns=())
        self.tree.configure(columns=())

        # 列名加 "c_" 前缀，避免与 Tk 内部列名冲突
        safe_cols = tuple(f"c_{i}" for i in range(len(columns)))
        self.tree.configure(columns=safe_cols)

        # 设置显示列（隐藏前N列）
        display_indices = list(range(hide_first_cols, len(columns)))
        self.tree.configure(displaycolumns=tuple(safe_cols[i] for i in display_indices))

        # 设置列标题
        for i, col in enumerate(columns):
            self.tree.heading(safe_cols[i], text=str(col))
            self.tree.column(safe_cols[i], width=100, minwidth=50)

        # 配置交替行颜色（模拟网格线效果）
        self.tree.tag_configure("evenrow", background="#f0f0f0")
        self.tree.tag_configure("oddrow", background="#ffffff")

        # 填充数据（交替行颜色）
        for idx, row in enumerate(rows):
            values = [str(v) if v is not None else "" for v in row]
            tag = "evenrow" if idx % 2 == 0 else "oddrow"
            self.tree.insert("", tk.END, values=values, tags=(tag,))

    def _on_cell_changed(self, event):
        """选中行变化时加载属性"""
        selected = self.combo_type.get()
        if not selected:
            return

        selection = self.tree.selection()
        if not selection:
            return

        item = self.tree.selection()[0]
        item_index = self.tree.index(item)
        self._load_properties_for_row(item_index, selected, self.current_columns)

    def _load_properties_for_row(self, row_index, selected_type, columns):
        """加载选中行的属性详情"""
        if row_index < 0 or row_index >= len(self.current_rows):
            return

        row_data = self.current_rows[row_index]

        self.prop_tree.delete(*self.prop_tree.get_children())

        if selected_type == "COUNT":
            file_name = row_data[0] if row_data else None
            if file_name is not None:
                self._load_pipeline_specs(file_name)
        else:
            file_name = row_data[0] if len(row_data) > 0 else None
            uci = row_data[1] if len(row_data) > 1 else None
            if file_name is not None and uci is not None:
                self._load_point_data(file_name, uci)

    def _load_pipeline_specs(self, file_name):
        """加载 PIPELINE 属性到属性面板"""
        columns, rows = self.db_helper.fill_data_table(
            "SELECT PIPINGSPEC, INSULATIONSPEC, PAINTINGSPEC, TRACINGSPEC, REFERENCE "
            "FROM PIPELINE WHERE FILENAME=?", (file_name,))

        if not rows:
            return

        row = rows[0]
        # 基本属性分类
        cat_basic = self.prop_tree.insert("", tk.END, text="基本信息", values=("",), open=True)
        self.prop_tree.insert(cat_basic, tk.END, text="文件名", values=(str(file_name),))
        self.prop_tree.insert(cat_basic, tk.END, text="管线编号", values=(self._safe_str(row[4]),))
        self.prop_tree.insert(cat_basic, tk.END, text="管道等级", values=(self._safe_str(row[0]),))
        self.prop_tree.insert(cat_basic, tk.END, text="保温等级", values=(self._safe_str(row[1]),))
        self.prop_tree.insert(cat_basic, tk.END, text="油漆等级", values=(self._safe_str(row[2]),))
        self.prop_tree.insert(cat_basic, tk.END, text="伴热等级", values=(self._safe_str(row[3]),))

        # 动态属性
        attr_cols, attr_rows = self.db_helper.fill_data_table(
            "SELECT ATTRIBUTE_NAME, ATTRIBUTE_VALUE FROM COMPONENT_ATTRIBUTES "
            "WHERE FILENAME=? AND UCI LIKE 'PIPELINE:%' ORDER BY ATTRIBUTE_NAME", (file_name,))

        if attr_rows:
            cat_attr = self.prop_tree.insert("", tk.END, text="属性", values=("",), open=True)
            for ar in attr_rows:
                attr_name = self._safe_str(ar[0])
                attr_value = self._safe_str(ar[1])
                self.prop_tree.insert(cat_attr, tk.END, text=attr_name, values=(attr_value,))

    def _load_point_data(self, file_name, uci):
        """加载坐标点数据到属性面板"""
        params = (uci, file_name)

        # END_POINT
        self._add_points_to_tree("END_POINT", "端点", "END_POINTX", "END_POINTY", "END_POINTZ", params)
        # CENTRE_POINT
        self._add_points_to_tree("CENTRE_POINT", "中心点", "CENTRE_POINTX", "CENTRE_POINTY", "CENTRE_POINTZ", params)
        # BRANCH1_POINT
        self._add_points_to_tree("BRANCH1_POINT", "分支点", "BRANCH1_POINTX", "BRANCH1_POINTY", "BRANCH1_POINTZ", params)
        # JACKET_POINT
        self._add_points_to_tree("JACKET_POINT", "夹套点", "JACKET_POINTX", "JACKET_POINTY", "JACKET_POINTZ", params)
        # CO_ORDS
        self._add_points_to_tree("CO_ORDS", "坐标点", "CO_ORDSX", "CO_ORDSY", "CO_ORDSZ", params)

    def _add_points_to_tree(self, table_name, display_name, col_x, col_y, col_z, params):
        """查询坐标点表并添加到属性树"""
        try:
            columns, rows = self.db_helper.fill_data_table(
                f"SELECT {col_x}, {col_y}, {col_z}, NPD, TYPE FROM {table_name} WHERE UCI IS ? AND FILENAME=?",
                params)
        except Exception:
            return

        if not rows:
            return

        for i, row in enumerate(rows):
            label = display_name if i == 0 else f"{display_name}{i + 1}"
            parent = self.prop_tree.insert("", tk.END, text=label, values=("",), open=True)
            self.prop_tree.insert(parent, tk.END, text="X坐标", values=(self._safe_str(row[0]),))
            self.prop_tree.insert(parent, tk.END, text="Y坐标", values=(self._safe_str(row[1]),))
            self.prop_tree.insert(parent, tk.END, text="Z坐标", values=(self._safe_str(row[2]),))
            self.prop_tree.insert(parent, tk.END, text="公称直径", values=(self._safe_str(row[3]),))
            self.prop_tree.insert(parent, tk.END, text="端面形式", values=(self._safe_str(row[4]),))

    # ==================== 导出 Excel ====================

    def _on_export(self):
        """导出数据到 Excel"""
        if not HAS_OPENPYXL:
            messagebox.showerror("错误", "缺少 openpyxl 库，请运行: pip install openpyxl")
            return

        valid_name = re.sub(r'[\x00-\x1F\x7F<>?*:"/\\|]', '', os.path.splitext(os.path.basename(self.db_path))[0])
        valid_name = valid_name.rstrip(' .-').lstrip('.')
        default_file = f"{valid_name}.xlsx"
        output_path = filedialog.asksaveasfilename(
            title="导出到 Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=default_file)
        if not output_path:
            return

        try:
            wb = Workbook()
            # 删除默认 sheet
            wb.remove(wb.active)

            # ========== PIPELINE 表 ==========
            # 获取所有属性名
            attr_cols, attr_rows = self.db_helper.fill_data_table(
                "SELECT DISTINCT ATTRIBUTE_NAME FROM COMPONENT_ATTRIBUTES WHERE UCI LIKE 'PIPELINE:%' ORDER BY ATTRIBUTE_NAME")
            attribute_names = [r[0] for r in attr_rows]

            pipe_cols, pipe_rows = self.db_helper.fill_data_table(
                "SELECT p.FILENAME, p.REFERENCE, p.PIPINGSPEC, p.INSULATIONSPEC, p.PAINTINGSPEC, p.TRACINGSPEC FROM PIPELINE p")

            if pipe_rows:
                ws = wb.create_sheet("PIPELINE")
                # 列标题（移除 FILENAME）
                headers = ["REFERENCE", "PIPINGSPEC", "INSULATIONSPEC", "PAINTINGSPEC", "TRACINGSPEC"]
                headers.extend(attribute_names)
                for j, h in enumerate(headers):
                    ws.cell(row=1, column=j + 1, value=h)

                for i, prow in enumerate(pipe_rows):
                    ws.cell(row=i + 2, column=1, value=self._safe_str(prow[1]))
                    ws.cell(row=i + 2, column=2, value=self._safe_str(prow[2]))
                    ws.cell(row=i + 2, column=3, value=self._safe_str(prow[3]))
                    ws.cell(row=i + 2, column=4, value=self._safe_str(prow[4]))
                    ws.cell(row=i + 2, column=5, value=self._safe_str(prow[5]))

                    # 获取属性
                    pipeline_uci = f"PIPELINE:{self._safe_str(prow[1])}"
                    for j, attr_name in enumerate(attribute_names):
                        attr_val_cols, attr_val_rows = self.db_helper.fill_data_table(
                            "SELECT ATTRIBUTE_VALUE FROM COMPONENT_ATTRIBUTES WHERE UCI=? AND ATTRIBUTE_NAME=?",
                            (pipeline_uci, attr_name))
                        val = self._safe_str(attr_val_rows[0][0]) if attr_val_rows else ""
                        ws.cell(row=i + 2, column=6 + j, value=val)

            # ========== 组件表 ==========
            for comp_type in self.COMPONENT_TYPES:
                view_name = "X_UNION" if comp_type == "_UNION" else f"X{comp_type}"
                cols, rows = self.db_helper.fill_data_table(f"SELECT * FROM {view_name}")
                if rows and "FILENAME" in cols:
                    ws = wb.create_sheet(comp_type)
                    # 移除 FILENAME 列
                    file_idx = cols.index("FILENAME")
                    display_cols = [c for i, c in enumerate(cols) if i != file_idx]

                    # 列标题
                    for j, h in enumerate(display_cols):
                        ws.cell(row=1, column=j + 1, value=h)

                    # 数据行
                    for i, drow in enumerate(rows):
                        for j, val in enumerate(drow):
                            if j == file_idx:
                                continue
                            col_idx = j if j < file_idx else j - 1
                            ws.cell(row=i + 2, column=col_idx + 1, value=self._safe_str(val))

            # 保存文件
            wb.save(output_path)
            messagebox.showinfo("成功", "成功输出文本到Excel！")

        except Exception as e:
            messagebox.showerror("错误", str(e))

    # ==================== 辅助方法 ====================

    @staticmethod
    def _safe_str(value):
        if value is None:
            return ""
        return str(value)


def main():
    root = ttkb.Window(themename="cosmo")
    app = PcfApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
